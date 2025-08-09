from pathlib import Path
import pandas as pd
from typing import List
import PyPDF2
import openpyxl
from llama_index.core import (
    Document
)
from llama_index.core.node_parser import SentenceSplitter
from llama_index.core.schema import TextNode

import logging
# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class DocumentProcessor:
    def __init__(self):
        self.node_parser = SentenceSplitter(
            chunk_size=1024,  # Increased from 512 to 1024
            chunk_overlap=100  # Increased proportionally
        )

    def extract_text_from_pdf(self, pdf_path: str, filename: str) -> List[Document]:
        """Extract text from PDF file"""
        documents = []
        try:
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                total_pages = len(pdf_reader.pages)
                logger.info(f"Processing PDF {filename} with {total_pages} pages")

                for page_num, page in enumerate(pdf_reader.pages, 1):
                    text = page.extract_text()
                    if text.strip():
                        doc = Document(
                            text=text,
                            metadata={
                                "file_name": filename,
                                "page_number": page_num,
                                "total_pages": total_pages,
                                "source": f"{filename}_page_{page_num}",
                                "document_type": "pdf"
                            }
                        )
                        documents.append(doc)
                    else:
                        logger.warning(f"Empty page {page_num} in {filename}")
                        
        except Exception as e:
            logger.error(f"Error processing PDF {filename}: {str(e)}")
            raise

        return documents

    def load_qa_from_csv(self, csv_path: str) -> List[Document]:
        """Load Q&A pairs from CSV file for both exact and semantic matching"""
        documents = []
        try:
            df = pd.read_csv(csv_path)
            logger.info(f"Processing CSV with {len(df)} rows")
            
            # Handle different possible column names
            question_col = None
            answer_col = None
            
            # Check column A and B first (as mentioned by user)
            if len(df.columns) >= 2:
                # Try columns A and B by index
                question_col = df.columns[0]  # Column A
                answer_col = df.columns[1]    # Column B
                logger.info(f"Using columns A and B: {question_col} -> {answer_col}")
            
            # Also check by name as fallback
            if not question_col or not answer_col:
                for col in df.columns:
                    col_lower = col.lower().strip()
                    if any(keyword in col_lower for keyword in ['question', 'questions', 'q', 'سؤال', 'استفسار']):
                        question_col = col
                    elif any(keyword in col_lower for keyword in ['answer', 'answers', 'a', 'response', 'جواب', 'إجابة']):
                        answer_col = col
            
            if question_col and answer_col:
                logger.info(f"Using Q&A columns: {question_col} -> {answer_col}")
                return self._process_qa_data(df, question_col, answer_col, csv_path, 'csv')
            else:
                logger.warning(f"Could not find question/answer columns in CSV. Available columns: {df.columns.tolist()}")
                
        except Exception as e:
            logger.error(f"Error processing CSV: {str(e)}")
            raise

        return documents

    def load_qa_from_excel(self, excel_path: str) -> List[Document]:
        """Load Q&A pairs from Excel file for both exact and semantic matching"""
        documents = []
        try:
            workbook = openpyxl.load_workbook(excel_path, read_only=True)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                logger.info(f"Processing Excel sheet: {sheet_name}")
                
                # Convert sheet to DataFrame for easier processing
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(list(row) if row else [])
                
                if len(data) < 1:
                    logger.warning(f"Sheet {sheet_name} is empty")
                    continue
                
                # Create DataFrame from Excel data
                df = pd.DataFrame(data)
                
                # Use columns A and B (index 0 and 1)
                if len(df.columns) >= 2:
                    # Remove header row if it looks like headers
                    if len(df) > 0:
                        first_row = df.iloc[0]
                        if (str(first_row[0]).lower().strip() in ['question', 'q', 'سؤال'] or 
                            str(first_row[1]).lower().strip() in ['answer', 'a', 'جواب', 'إجابة']):
                            df = df.iloc[1:].reset_index(drop=True)
                    
                    documents = self._process_qa_data(df, 0, 1, excel_path, 'excel', sheet_name)
                    
                    if documents:
                        logger.info(f"Successfully loaded {len(documents)} Q&A pairs from sheet {sheet_name}")
                        break  # Found Q&A data, no need to check other sheets
                    
        except Exception as e:
            logger.error(f"Error processing Excel file: {str(e)}")
            raise

        return documents

    def _process_qa_data(self, df, question_col, answer_col, file_path, doc_type, sheet_name=None):
        """Common method to process Q&A data from CSV/Excel with optimized metadata"""
        documents = []
        
        for idx, row in df.iterrows():
            question = row[question_col]
            answer = row[answer_col]
            
            if pd.notna(question) and pd.notna(answer) and str(question).strip() and str(answer).strip():
                question_clean = str(question).strip()
                answer_clean = str(answer).strip()
                
                # Skip obviously invalid data
                if len(question_clean) < 3 or len(answer_clean) < 3:
                    continue

                # Create document for semantic search
                qa_text = f"Question: {question_clean}\nAnswer: {answer_clean}"
                
                # Optimized metadata with shorter keys and values
                metadata = {
                    "file_name": Path(file_path).name,
                    "page_number": idx + 1,
                    "source": f"{doc_type}_row_{idx + 1}",
                    "type": "qa_pair",
                    "document_type": doc_type,
                    "original_question": question_clean,
                    "original_answer": answer_clean
                }
                
                # Only add sheet_name if it exists and is short
                if sheet_name and len(str(sheet_name)) < 50:
                    metadata["sheet_name"] = str(sheet_name)
                
                doc = Document(text=qa_text, metadata=metadata)
                documents.append(doc)
                
        return documents

    def create_nodes_with_metadata(self, all_documents: List[Document]) -> List[TextNode]:
        """Create nodes from documents using the node parser with metadata optimization"""
        nodes = []
        for doc in all_documents:
            try:
                # Create a copy with optimized metadata
                logger.info(f"text are {doc.text}")
                
                # Create a new document with optimized metadata
                optimized_doc = Document(
                    text=doc.text,
                    metadata=doc.metadata
                )
                
                doc_nodes = self.node_parser.get_nodes_from_documents([optimized_doc])
                logger.info(f"doc nodes are {doc_nodes}")
                nodes.extend(doc_nodes)
                
            except Exception as e:
                logger.warning(f"Error processing document with metadata {doc.metadata.get('source', 'unknown')}: {str(e)}")
                # Try with minimal metadata as fallback
                try:
                    minimal_doc = Document(
                        text=doc.text,
                        metadata={
                            "file_name": str(doc.metadata.get('file_name', 'unknown'))[:50],
                            "page_number": doc.metadata.get('page_number', 1),
                            "source": str(doc.metadata.get('source', 'unknown'))[:30]
                        }
                    )
                    doc_nodes = self.node_parser.get_nodes_from_documents([minimal_doc])
                    nodes.extend(doc_nodes)
                    logger.info(f"Successfully processed document with minimal metadata")
                except Exception as e2:
                    logger.error(f"Failed to process document even with minimal metadata: {str(e2)}")
                    continue
                
        return nodes
    
    def _optimize_metadata(self, metadata: dict) -> dict:
        """Optimize metadata to prevent size issues"""
        optimized = {}
        
        # Define maximum lengths for each field
        field_limits = {
            'file_name': 100,
            'original_question': 2000,
            'original_answer': 20000,
            'source': 50,
            'sheet_name': 30,
            'document_type': 20
        }
        
        for key, value in metadata.items():
            if key in field_limits and isinstance(value, str):
                max_length = field_limits[key]
                if len(value) > max_length:
                    optimized[key] = value[:max_length] + "..."
                else:
                    optimized[key] = value
            elif key == 'page_number':
                optimized[key] = int(value) if value else 1
            elif isinstance(value, str) and len(value) > 100:
                # Generic string truncation
                optimized[key] = value[:100] + "..."
            else:
                optimized[key] = value
                
        return optimized

